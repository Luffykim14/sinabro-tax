"""
국세청 전자세금계산서 일괄발급 엑셀 양식 생성 모듈
실제 국세청 양식 템플릿에 직접 데이터를 입력하는 방식
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
import shutil
from datetime import datetime

# ===== 공급자 고정 정보 (시나브로마케팅) =====
SUPPLIER = {
    "type_code": "01",
    "registration_number": "7111802350",
    "company_name": "시나브로마케팅",
    "representative": "김수현",
    "address": "서울특별시 강동구 천호대로 1037, 9층 b23호 (천호동, 세우빌딩)",
    "business_type": "서비스업",
    "business_item": "광고대행업",
    "email": "info@sinabro.biz",
    "payment_type": "01"
}

# 템플릿 파일 경로 (프로그램 폴더에 저장)
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "template.xlsx")


def create_excel(rows: list, output_folder: str) -> str:
    """
    국세청 양식 템플릿에 데이터 직접 입력하여 엑셀 생성
    """
    # 템플릿 있으면 사용, 없으면 기본 생성
    if os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb.worksheets[0]
        # 기존 데이터 행 초기화 (7행부터)
        for row_idx in range(7, ws.max_row + 1):
            for col_idx in range(1, 60):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value not in (None, 0, ""):
                    # 수식이 아닌 셀만 초기화
                    if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                        cell.value = None
    else:
        wb = _create_basic_template()
        ws = wb.active

    # ===== 데이터 입력 (7행부터) =====
    for row_idx, row_data in enumerate(rows, 7):
        buyer = row_data.get('supplier_info', {})
        totals = row_data.get('totals', {})
        items = row_data.get('items', [])
        date_info = row_data.get('date_info', {})

        full_date = date_info.get('full_date', datetime.now().strftime('%Y%m%d'))
        day = date_info.get('day', '01')

        item1 = items[0] if len(items) > 0 else {}
        item2 = items[1] if len(items) > 1 else {}

        # 공급자 정보 (col1~10)
        ws.cell(row=row_idx, column=1).value = SUPPLIER["type_code"]
        ws.cell(row=row_idx, column=2).value = int(full_date)
        ws.cell(row=row_idx, column=3).value = SUPPLIER["registration_number"]
        ws.cell(row=row_idx, column=4).value = ""  # 종사업장번호
        ws.cell(row=row_idx, column=5).value = SUPPLIER["company_name"]
        ws.cell(row=row_idx, column=6).value = SUPPLIER["representative"]
        ws.cell(row=row_idx, column=7).value = SUPPLIER["address"]
        ws.cell(row=row_idx, column=8).value = SUPPLIER["business_type"]
        ws.cell(row=row_idx, column=9).value = SUPPLIER["business_item"]
        ws.cell(row=row_idx, column=10).value = SUPPLIER["email"]

        # 공급받는자 정보 (col11~19)
        ws.cell(row=row_idx, column=11).value = buyer.get("registration_number", "")
        ws.cell(row=row_idx, column=12).value = ""  # 종사업장번호
        ws.cell(row=row_idx, column=13).value = buyer.get("company_name", "")
        ws.cell(row=row_idx, column=14).value = buyer.get("representative", "")
        ws.cell(row=row_idx, column=15).value = buyer.get("address", "")
        ws.cell(row=row_idx, column=16).value = buyer.get("business_type", "")
        ws.cell(row=row_idx, column=17).value = buyer.get("business_item", "")
        ws.cell(row=row_idx, column=18).value = buyer.get("email", "")
        ws.cell(row=row_idx, column=19).value = ""  # 이메일2

        # 합계 (col20~22)
        ws.cell(row=row_idx, column=20).value = totals.get("supply_amount", 0)
        ws.cell(row=row_idx, column=21).value = totals.get("tax_amount", 0)
        ws.cell(row=row_idx, column=22).value = ""  # 비고

        # 품목1 (col23~30)
        ws.cell(row=row_idx, column=23).value = day           # 일자1
        ws.cell(row=row_idx, column=24).value = item1.get("name", "")
        ws.cell(row=row_idx, column=25).value = ""            # 규격1
        ws.cell(row=row_idx, column=26).value = item1.get("quantity", "")
        ws.cell(row=row_idx, column=27).value = item1.get("unit_price", "")
        ws.cell(row=row_idx, column=28).value = item1.get("supply_amount", "")
        ws.cell(row=row_idx, column=29).value = item1.get("tax_amount", "")
        ws.cell(row=row_idx, column=30).value = ""            # 품목비고1

        # 품목2 (col31~38)
        ws.cell(row=row_idx, column=31).value = day           # 일자2
        ws.cell(row=row_idx, column=32).value = item2.get("name", "")
        ws.cell(row=row_idx, column=33).value = ""            # 규격2
        ws.cell(row=row_idx, column=34).value = item2.get("quantity", "")
        ws.cell(row=row_idx, column=35).value = item2.get("unit_price", "")
        ws.cell(row=row_idx, column=36).value = item2.get("supply_amount", "")
        ws.cell(row=row_idx, column=37).value = item2.get("tax_amount", "")
        ws.cell(row=row_idx, column=38).value = ""            # 품목비고2

        # 영수/청구 (마지막 컬럼 - 실제 양식에서 확인된 위치)
        # 영수청구는 별도 컬럼에 입력 (템플릿에 따라 위치가 다를 수 있음)

    # 파일 저장
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"세금계산서_업로드양식_{timestamp}.xlsx"
    filepath = os.path.join(output_folder, filename)
    wb.save(filepath)
    return filepath


def _create_basic_template():
    """템플릿 파일 없을 때 기본 양식 생성"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "시트1"

    orange_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        "전자(세금)계산서\n종류코드", "작성일자", "공급자\n등록번호", "공급자\n종사업장번호",
        "공급자\n상호", "공급자\n성명", "공급자\n사업장주소", "공급자\n업태", "공급자\n종목", "공급자\n이메일",
        "공급받는자\n등록번호", "공급받는자\n종사업장번호", "공급받는자\n상호", "공급받는자\n성명",
        "공급받는자\n사업장주소", "공급받는자\n업태", "공급받는자\n종목", "공급받는자\n이메일1", "공급받는자\n이메일2",
        "공급가액\n합계", "세액\n합계", "비고",
        "일자1", "품목1", "규격1", "수량1", "단가1", "공급가액1", "세액1", "품목비고1",
        "일자2", "품목2", "규격2", "수량2", "단가2", "공급가액2", "세액2", "품목비고2",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_idx, value=header)
        cell.fill = orange_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[6].height = 40
    return wb


def save_template(template_file_path: str):
    """업로드된 국세청 양식을 템플릿으로 저장"""
    shutil.copy2(template_file_path, TEMPLATE_PATH)
